from tkinter import *
from openpyxl import load_workbook

# Load the workbook and select the active sheet
wb = load_workbook('path_to_your_excel_file.xlsx')
sheet = wb.active

def excel_setup():
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # Set column headers
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

def clear_fields():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert_data():
    if (name_field.get() == "" or course_field.get() == "" or sem_field.get() == "" or form_no_field.get() == "" or contact_no_field.get() == "" or email_id_field.get() == "" or address_field.get() == ""):
        print("Please fill in all fields")
    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        wb.save('path_to_your_excel_file.xlsx')
        clear_fields()

# Set up the GUI
root = Tk()
root.configure(background='light green')
root.title("Registration Form")
root.geometry("500x300")

excel_setup()

# Create and place labels and entry fields
Label(root, text="Name", bg="light green").grid(row=0, column=0)
name_field = Entry(root)
name_field.grid(row=0, column=1)

Label(root, text="Course", bg="light green").grid(row=1, column=0)
course_field = Entry(root)
course_field.grid(row=1, column=1)

Label(root, text="Semester", bg="light green").grid(row=2, column=0)
sem_field = Entry(root)
sem_field.grid(row=2, column=1)

Label(root, text="Form Number", bg="light green").grid(row=3, column=0)
form_no_field = Entry(root)
form_no_field.grid(row=3, column=1)

Label(root, text="Contact Number", bg="light green").grid(row=4, column=0)
contact_no_field = Entry(root)
contact_no_field.grid(row=4, column=1)

Label(root, text="Email id", bg="light green").grid(row=5, column=0)
email_id_field = Entry(root)
email_id_field.grid(row=5, column=1)

Label(root, text="Address", bg="light green").grid(row=6, column=0)
address_field = Entry(root)
address_field.grid(row=6, column=1)

# Create and place buttons
Button(root, text="Submit", command=insert_data).grid(row=7, column=1)
Button(root, text="Clear", command=clear_fields).grid(row=7, column=0)

root.mainloop()
